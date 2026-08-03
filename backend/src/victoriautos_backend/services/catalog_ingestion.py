"""Helpers for building a Colombian vehicle catalog from permitted public sources."""

from __future__ import annotations

import csv
import re
import unicodedata
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CATALOG_COLUMNS = (
    "source",
    "source_dataset_id",
    "source_record_id",
    "source_updated_at",
    "vehicle_type",
    "brand",
    "line",
    "version",
    "model_year",
    "engine_cc",
    "market_value_cop",
    "observations",
)

FIELD_ALIASES = {
    "brand": ("marca", "fabricante", "marca_vehiculo", "marca_del_vehiculo"),
    "line": (
        "linea",
        "linea_vehiculo",
        "linea_del_vehiculo",
        "referencia",
        "referencia1",
        "referencia_1",
    ),
    "version": (
        "version",
        "version_vehiculo",
        "referencia2",
        "referencia_2",
        "referencia3",
        "referencia_3",
    ),
    "model_year": (
        "modelo",
        "modelo_vehiculo",
        "modelo_del_vehiculo",
        "modelo_del_automotor",
        "ano_modelo",
        "anio_modelo",
        "ano_del_modelo",
    ),
    "vehicle_type": (
        "clase",
        "clase_vehiculo",
        "clase_del_vehiculo",
        "tipo_vehiculo",
        "tipo_de_automotor",
        "categoria",
        "tipologia",
    ),
    "source_record_id": ("codigo", "codigo_fasecolda", "codigofasecolda"),
    "market_value_cop": ("valor", "valor_comercial", "valor_fasecolda"),
    "engine_cc": ("cilindraje", "cilindraje_cc", "cilindraje_del_motor"),
    "observations": ("cantidad", "conteo", "cantidad_vehiculos", "total_vehiculos"),
}
CATALOG_CONTEXT_TERMS = ("parque automotor", "vehiculos matriculados", "vehiculos registrados")
EXCLUDED_CONTEXT_TERMS = (
    "accidente",
    "hardware",
    "inventario",
    "mototaxi",
    "motocarro",
    "recoleccion",
    "repartidor",
    "residuo",
    "siniestro",
)


def normalize_header(value: Any) -> str:
    """Normalize a source heading for resilient semantic matching."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "_", text.casefold()).strip("_")


def normalize_catalog_text(value: Any) -> str:
    """Normalize catalog labels without inventing aliases or changing their meaning."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return re.sub(r"\s+", " ", text).upper()


def parse_model_year(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        year = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None
    maximum = datetime.now(tz=UTC).year + 1
    return year if 1900 <= year <= maximum else None


def find_semantic_fields(
    columns: Sequence[tuple[str, str]], *, required: Sequence[str] = ()
) -> dict[str, str]:
    """Map semantic catalog fields to Socrata field identifiers or sheet headings."""
    normalized = [
        (field_id, normalize_header(label) or normalize_header(field_id))
        for field_id, label in columns
    ]
    result: dict[str, str] = {}
    for semantic, aliases in FIELD_ALIASES.items():
        alias_set = set(aliases)
        match = next(
            (field_id for field_id, label in normalized if label in alias_set),
            None,
        )
        if match:
            result[semantic] = match
    missing = [field for field in required if field not in result]
    if missing:
        raise ValueError(f"Missing required catalog fields: {', '.join(missing)}")
    return result


def is_vehicle_catalog_dataset(result: Mapping[str, Any]) -> bool:
    """Reject look-alike datasets whose marca/modelo fields describe other assets."""
    resource = result.get("resource", {})
    context = normalize_header(f"{resource.get('name', '')} {resource.get('description', '')}")
    context = context.replace("_", " ")
    if any(term in context for term in EXCLUDED_CONTEXT_TERMS):
        return False
    return any(term in context for term in CATALOG_CONTEXT_TERMS)


def _find_workbook_header(rows: Sequence[Sequence[Any]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:50]):
        columns = [(str(column_index), value) for column_index, value in enumerate(row)]
        fields = find_semantic_fields(columns)
        if "brand" in fields and "line" in fields:
            return index, {name: int(column) for name, column in fields.items()}
    raise ValueError("Could not find a header row containing brand and line/reference columns")


def parse_fasecolda_workbook(path: Path) -> list[dict[str, Any]]:
    """Parse an authorized Fasecolda XLSX publication into normalized catalog rows.

    Both long-form sheets (one model year per row) and the traditional wide-form
    guide (one price column per model year) are supported.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            raw_rows = list(sheet.iter_rows(values_only=True))
            if not raw_rows:
                continue
            try:
                header_index, fields = _find_workbook_header(raw_rows)
            except ValueError:
                continue
            header = raw_rows[header_index]
            version_aliases = set(FIELD_ALIASES["version"])
            version_columns = [
                column_index
                for column_index, value in enumerate(header)
                if normalize_header(value) in version_aliases
            ]
            wide_years = {
                column_index: year
                for column_index, value in enumerate(header)
                if (year := parse_model_year(value)) is not None
            }
            for row in raw_rows[header_index + 1 :]:
                brand = normalize_catalog_text(_cell(row, fields.get("brand")))
                line = normalize_catalog_text(_cell(row, fields.get("line")))
                if not brand or not line:
                    continue
                base = {
                    "source": "fasecolda",
                    "source_dataset_id": path.name,
                    "source_record_id": normalize_catalog_text(
                        _cell(row, fields.get("source_record_id"))
                    ),
                    "source_updated_at": "",
                    "vehicle_type": normalize_catalog_text(_cell(row, fields.get("vehicle_type"))),
                    "brand": brand,
                    "line": line,
                    "version": " ".join(
                        value
                        for column_index in version_columns
                        if (value := normalize_catalog_text(_cell(row, column_index)))
                    ),
                    "engine_cc": _number_text(_cell(row, fields.get("engine_cc"))),
                    "observations": "",
                }
                long_year = parse_model_year(_cell(row, fields.get("model_year")))
                if long_year is not None:
                    output.append(
                        {
                            **base,
                            "model_year": long_year,
                            "market_value_cop": _number_text(
                                _cell(row, fields.get("market_value_cop"))
                            ),
                        }
                    )
                    continue
                for column_index, year in wide_years.items():
                    value = _cell(row, column_index)
                    if value not in (None, "", 0, "0"):
                        output.append(
                            {
                                **base,
                                "model_year": year,
                                "market_value_cop": _number_text(value),
                            }
                        )
    finally:
        workbook.close()
    return _deduplicate_rows(output)


def _cell(row: Sequence[Any], index: int | None) -> Any:
    return None if index is None or index >= len(row) else row[index]


def parse_mintransport_workbook(
    path: Path, *, source_dataset_id: str | None = None, source_updated_at: str = ""
) -> list[dict[str, Any]]:
    """Parse one official MinTransporte base-gravable XLSX table."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, row in enumerate(rows[:30])
                    if "marca" in {normalize_header(value).split("_")[0] for value in row}
                    and "linea" in {normalize_header(value).split("_")[0] for value in row}
                ),
                None,
            )
            if header_index is None:
                continue
            primary_header = rows[header_index]
            secondary_header = rows[header_index + 1] if header_index + 1 < len(rows) else ()
            headings = [
                normalize_header(primary or secondary)
                for primary, secondary in zip(primary_header, secondary_header, strict=False)
            ]
            brand_index = _heading_index(headings, "marca")
            line_index = _heading_index(headings, "linea")
            if brand_index is None or line_index is None:
                continue
            id_index = _heading_index(headings, "id")
            type_index = _heading_index(headings, "clase")
            engine_index = _heading_index(headings, "cilindraje")
            year_columns: dict[int, int] = {}
            for column_index, (primary, secondary) in enumerate(
                zip(primary_header, secondary_header, strict=False)
            ):
                label = f"{primary or ''} {secondary or ''}"
                if "anterior" in normalize_header(label):
                    continue
                if (year := parse_model_year(secondary or primary)) is not None:
                    year_columns[column_index] = year

            for row in rows[header_index + 2 :]:
                brand = normalize_catalog_text(_cell(row, brand_index))
                line = normalize_catalog_text(_cell(row, line_index))
                if not brand or not line:
                    continue
                base = {
                    "source": "mintransporte.gov.co",
                    "source_dataset_id": source_dataset_id or path.name,
                    "source_record_id": normalize_catalog_text(_cell(row, id_index)),
                    "source_updated_at": source_updated_at,
                    "vehicle_type": normalize_catalog_text(_cell(row, type_index)),
                    "brand": brand,
                    "line": line,
                    "version": "",
                    "engine_cc": _number_text(_cell(row, engine_index)),
                    "observations": "",
                }
                for column_index, year in year_columns.items():
                    value = _cell(row, column_index)
                    if value in (None, "", 0, "0"):
                        continue
                    value_text = _number_text(value)
                    try:
                        value_cop = str(int(float(value_text) * 1000))
                    except ValueError:
                        continue
                    output.append(
                        {
                            **base,
                            "model_year": year,
                            "market_value_cop": value_cop,
                        }
                    )
    finally:
        workbook.close()
    return _deduplicate_rows(output)


def _heading_index(headings: Sequence[str], prefix: str) -> int | None:
    return next(
        (index for index, heading in enumerate(headings) if heading.split("_")[0] == prefix),
        None,
    )


def _number_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        number = float(str(value).replace(",", "").strip())
    except ValueError:
        return normalize_catalog_text(value)
    return str(int(number)) if number.is_integer() else str(number)


def _deduplicate_rows(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    keyed: dict[tuple[Any, ...], dict[str, Any]] = {}
    key_fields = (
        "source",
        "source_dataset_id",
        "source_record_id",
        "vehicle_type",
        "brand",
        "line",
        "version",
        "model_year",
    )
    for row in rows:
        normalized = {column: row.get(column, "") for column in CATALOG_COLUMNS}
        keyed[tuple(normalized[field] for field in key_fields)] = normalized
    return list(keyed.values())


def merge_catalog_rows(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Merge source rows for form use while retaining coverage and observation counts."""
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    sources: dict[tuple[Any, ...], set[str]] = defaultdict(set)
    fields = ("vehicle_type", "brand", "line", "version", "model_year", "engine_cc")
    for row in rows:
        key = tuple(row.get(field, "") for field in fields)
        sources[key].add(str(row.get("source_dataset_id", "")))
        observations = _safe_int(row.get("observations"))
        if key not in grouped:
            grouped[key] = {field: row.get(field, "") for field in fields}
            grouped[key]["observations"] = 0
        grouped[key]["observations"] += observations
    return [
        {
            **grouped[key],
            "source_count": len(source_ids),
            "source_dataset_ids": ";".join(sorted(source_ids)),
        }
        for key, source_ids in sources.items()
    ]


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def write_catalog_csv(path: Path, rows: Iterable[Mapping[str, Any]]) -> int:
    materialized = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=CATALOG_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(materialized)
    return len(materialized)


def write_merged_csv(path: Path, rows: Iterable[Mapping[str, Any]]) -> int:
    materialized = list(rows)
    fields = (
        "vehicle_type",
        "brand",
        "line",
        "version",
        "model_year",
        "engine_cc",
        "observations",
        "source_count",
        "source_dataset_ids",
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(materialized)
    return len(materialized)
