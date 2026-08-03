from pathlib import Path

from openpyxl import Workbook

from victoriautos_backend.services.catalog_ingestion import (
    find_semantic_fields,
    is_vehicle_catalog_dataset,
    merge_catalog_rows,
    normalize_catalog_text,
    parse_fasecolda_csv,
    parse_fasecolda_workbook,
    parse_mintransport_workbook,
    parse_model_year,
)


def test_normalizes_catalog_values_without_losing_accents() -> None:
    assert normalize_catalog_text("  Mercedes-Benz   Clase A ") == "MERCEDES-BENZ CLASE A"
    assert normalize_catalog_text("Citroën") == "CITROËN"
    assert parse_model_year("2022.0") == 2022
    assert parse_model_year("1880") is None


def test_detects_spanish_catalog_fields() -> None:
    fields = find_semantic_fields(
        [
            ("marca_del_vehiculo", "MARCA DEL VEHÍCULO"),
            ("linea", "LÍNEA"),
            ("modelo", "AÑO MODELO"),
            ("clase", "CLASE"),
            ("cantidad", "CANTIDAD"),
        ],
        required=("brand", "line", "model_year"),
    )

    assert fields == {
        "brand": "marca_del_vehiculo",
        "line": "linea",
        "model_year": "modelo",
        "vehicle_type": "clase",
        "observations": "cantidad",
    }


def test_parses_wide_fasecolda_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "guia.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Guía Fasecolda"])
    sheet.append(
        ["Código", "Marca", "Clase", "Referencia1", "Referencia2", "Referencia3", 2021, 2022]
    )
    sheet.append(["00123456", "Mazda", "Automovil", "3", "Touring", "2.0 AT", 60_000, 65_000])
    workbook.save(workbook_path)

    rows = parse_fasecolda_workbook(workbook_path)

    assert len(rows) == 2
    assert rows[0]["source_record_id"] == "00123456"
    assert rows[0]["brand"] == "MAZDA"
    assert rows[0]["line"] == "3"
    assert rows[0]["version"] == "TOURING 2.0 AT"
    assert rows[0]["model_year"] == 2021
    assert rows[0]["market_value_cop"] == "60000"


def test_parses_long_fasecolda_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "guia-larga.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Marca", "Línea", "Versión", "Modelo", "Valor comercial"])
    sheet.append(["Renault", "Duster", "Intens", 2020, 54_000_000])
    workbook.save(workbook_path)

    rows = parse_fasecolda_workbook(workbook_path)

    assert rows == [
        {
            "source": "fasecolda",
            "source_dataset_id": "guia-larga.xlsx",
            "source_record_id": "",
            "source_updated_at": "",
            "vehicle_type": "",
            "brand": "RENAULT",
            "line": "DUSTER",
            "version": "INTENS",
            "model_year": 2020,
            "engine_cc": "",
            "market_value_cop": "54000000",
            "observations": "",
        }
    ]


def test_parses_historical_fasecolda_csv(tmp_path: Path) -> None:
    csv_path = tmp_path / "fasecolda.csv"
    csv_path.write_text(
        "Marca,Clase,Codigo,Referencia1,Referencia2,Referencia3,2017,2018,Cilindraje\n"
        'Mazda,Automovil,02901001,3,Touring,"AT 2000CC",60000,65000,1998\n',
        encoding="utf-8",
    )

    rows = parse_fasecolda_csv(
        csv_path,
        source_dataset_id="fasecolda-2017",
        source_updated_at="2017-08-09",
    )

    assert len(rows) == 2
    assert rows[0] == {
        "source": "fasecolda-historical",
        "source_dataset_id": "fasecolda-2017",
        "source_record_id": "02901001",
        "source_updated_at": "2017-08-09",
        "vehicle_type": "AUTOMOVIL",
        "brand": "MAZDA",
        "line": "3",
        "version": "TOURING AT 2000CC",
        "model_year": 2017,
        "engine_cc": "1998",
        "market_value_cop": "60000000",
        "observations": "",
    }


def test_merges_sources_without_discarding_provenance() -> None:
    merged = merge_catalog_rows(
        [
            {
                "source_dataset_id": "aaaa-bbbb",
                "vehicle_type": "AUTOMOVIL",
                "brand": "MAZDA",
                "line": "3",
                "version": "",
                "model_year": 2020,
                "engine_cc": "2000",
                "observations": "5",
            },
            {
                "source_dataset_id": "cccc-dddd",
                "vehicle_type": "AUTOMOVIL",
                "brand": "MAZDA",
                "line": "3",
                "version": "",
                "model_year": 2020,
                "engine_cc": "2000",
                "observations": "7",
            },
        ]
    )

    assert merged == [
        {
            "vehicle_type": "AUTOMOVIL",
            "brand": "MAZDA",
            "line": "3",
            "version": "",
            "model_year": 2020,
            "engine_cc": "2000",
            "observations": 12,
            "source_count": 2,
            "source_dataset_ids": "aaaa-bbbb;cccc-dddd",
        }
    ]


def test_filters_non_catalog_datasets_even_when_columns_look_compatible() -> None:
    valid = {
        "resource": {
            "name": "Parque Automotor matriculados en Fusagasugá",
            "description": "Vehículos registrados en RUNT",
        }
    }
    unrelated = {
        "resource": {
            "name": "Rutas recolección residuos",
            "description": "Inventario de vehículos asignados",
        }
    }

    assert is_vehicle_catalog_dataset(valid)
    assert not is_vehicle_catalog_dataset(unrelated)


def test_parses_mintransport_two_row_headers_and_skips_historical_bucket(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "tabla-1.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["MINISTERIO DE TRANSPORTE"])
    sheet.append(["TABLA 1 - BASE GRAVABLE 2026"])
    sheet.append([])
    sheet.append(
        [
            "#",
            "ID",
            "TIPO (1)",
            "CLASE (2)",
            "MARCA (3)",
            "LINEA (4)",
            "CILINDRAJE (5)",
            "AÑO MODELO (7)",
            None,
        ]
    )
    sheet.append(["#", None, None, None, None, None, None, "2001 y Anteriores", 2022])
    sheet.append(
        [1, 123, "AUTOMOVILES", "AUTOMOVIL", "Mazda", "3 Touring AT", 2000, 12_000, 65_000]
    )
    workbook.save(workbook_path)

    rows = parse_mintransport_workbook(
        workbook_path,
        source_dataset_id="2026:tabla-1",
        source_updated_at="2026-01-06",
    )

    assert rows == [
        {
            "source": "mintransporte.gov.co",
            "source_dataset_id": "2026:tabla-1",
            "source_record_id": "123",
            "source_updated_at": "2026-01-06",
            "vehicle_type": "AUTOMOVIL",
            "brand": "MAZDA",
            "line": "3 TOURING AT",
            "version": "",
            "model_year": 2022,
            "engine_cc": "2000",
            "market_value_cop": "65000000",
            "observations": "",
        }
    ]
